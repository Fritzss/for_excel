from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from os import getcwd as pwd, path
import configparser
import chardet
from copy import copy
from tqdm import tqdm  # Импорт библиотеки для прогресс-бара


def detect_encoding(fp):
    with open(fp, 'rb') as f:
        det = chardet.universaldetector.UniversalDetector()
        for line in f:
            det.feed(line)
            if det.done:
                break
        det.close()
        return det.result['encoding']


def get_first_group_row(ws):
    """Находит первую строку группы"""
    for row in range(1, ws.max_row + 1):
        outline = ws.row_dimensions[row].outlineLevel
        if outline == 0 and row + 1 <= ws.max_row:
            next_outline = ws.row_dimensions[row + 1].outlineLevel
            if next_outline == 1:
                return row
    return 6  # Значение по умолчанию


def get_last_row(ws, tag):
    """Находит последнюю строку по тегу"""
    for row in range(ws.max_row, 0, -1):  # Оптимизировано: поиск снизу вверх
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and tag in str(cell_value):
            return row
    return ws.max_row


def get_groups(ws, tag, last_header_row):
    """Получает список групп с границами строк"""
    last_data_row = get_last_row(ws, tag)
    groups = []
    current_group = None

    for row in tqdm(range(last_header_row, last_data_row + 1), desc="Ищем группы"):

        if ws.row_dimensions[row].outlineLevel == 0 and ws.row_dimensions[row + 1].outlineLevel == 1:
            if current_group:
                current_group["last_row"] = row - 1
                groups.append(current_group)

            name = str(ws.cell(row=row, column=1).value or "").strip()
            current_group = {
                "filial": name,
                "first_row": row,
                "last_row": last_data_row
            }

    if current_group not in groups:
        current_group["last_row"] = last_data_row
        groups.append(current_group)

    return groups


def collapse_groups(ws, header_rows):
    # Скрываем все строки с уровнем группировки > 0
    for row in range(header_rows + 1, ws.max_row + 1):
        level = ws.row_dimensions[row].outline_level
        if level > 0:
            ws.row_dimensions[row].hidden = True

    # Оставляем видимыми заголовки
    for row in range(1, header_rows + 1):
        ws.row_dimensions[row].hidden = False

    # Настраиваем отображение группировки
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False


def create_group_sheets(wb, groups, source_sheet, last_header_row):
    """Создает листы для групп и копирует заголовки"""
    # Сбор информации об объединенных ячейках
    merged_ranges = [
        m for m in source_sheet.merged_cells.ranges
        if m.min_row < last_header_row
    ]
    # Кеширование стилей ячеек заголовка
    header_styles = {}
    print("Кеширование стилей заголовков...")
    for row in range(1, last_header_row):
        for col in range(1, source_sheet.max_column + 1):
            src_cell = source_sheet.cell(row=row, column=col)
            if src_cell.has_style:
                header_styles[(row, col)] = {
                    "font": copy(src_cell.font),
                    "border": copy(src_cell.border),
                    "fill": copy(src_cell.fill),
                    "number_format": src_cell.number_format,
                    "alignment": copy(src_cell.alignment)
                }

    # Создание листов с прогресс-баром
    for group in tqdm(groups, desc="Создание листов"):
        new_sheet = wb.create_sheet(title=group["filial"])
        new_sheet.freeze_panes = f'{get_column_letter(source_sheet.max_column + 1)}{last_header_row + 1}'
        # Копирование значений и стилей
        for row in range(1, last_header_row):
            # Копирование высоты строки
            new_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
            for col in range(1, source_sheet.max_column + 1):
                src_cell = source_sheet.cell(row=row, column=col)
                new_cell = new_sheet.cell(row=row, column=col, value=src_cell.value)

                if (row, col) in header_styles:
                    style = header_styles[(row, col)]
                    new_cell.font = style["font"]
                    new_cell.border = style["border"]
                    new_cell.fill = style["fill"]
                    new_cell.number_format = style["number_format"]
                    new_cell.alignment = style["alignment"]

        # Копирование ширины столбцов
        for col in range(1, source_sheet.max_column + 2):
            col_letter = get_column_letter(col)
            column_width = source_sheet.column_dimensions[col_letter].width
            new_sheet.column_dimensions[col_letter].width = column_width

        # Копирование объединенных ячеек
        for merged_range in merged_ranges:
            coord = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}:" \
                    f"{get_column_letter(merged_range.max_col)}{merged_range.max_row}"
            new_sheet.merge_cells(coord)



def copy_group_data(wb, source_sheet, groups, last_header_row):
    """Копирует данные групп на соответствующие листы"""
    src_ws = wb[source_sheet]

    # Прогресс-бар для групп
    for group in tqdm(groups, desc="Перенос данных"):
        new_ws = wb[group["filial"]]
        merged_ranges = [
            m for m in src_ws.merged_cells.ranges
            if m.min_row >= group["first_row"] and m.max_row <= group["last_row"]
        ]

        for row_idx in range(group["first_row"], group["last_row"] + 1):
            new_row_idx = row_idx - group["first_row"] + last_header_row
            new_ws.row_dimensions[new_row_idx].outline_level = src_ws.row_dimensions[row_idx].outline_level
            new_ws.row_dimensions[new_row_idx].height = src_ws.row_dimensions[row_idx].height
            # Копирование высоты строки
            new_ws.row_dimensions[new_row_idx].height = src_ws.row_dimensions[row_idx].height

            for col_idx in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(row=row_idx, column=col_idx)
                new_cell = new_ws.cell(row=new_row_idx, column=col_idx, value=src_cell.value)

                if src_cell.has_style:
                    new_cell.font = copy(src_cell.font)
                    new_cell.border = copy(src_cell.border)
                    new_cell.fill = copy(src_cell.fill)
                    new_cell.number_format = src_cell.number_format
                    new_cell.alignment = copy(src_cell.alignment)

            # Копирование ширины столбцов

            # for col_idx in range(1, src_ws.max_column + 1):
            #     col_letter = get_column_letter(col_idx)
            #     new_ws.column_dimensions[col_letter].width = \
            #         src_ws.column_dimensions[col_letter].width

        # merged cells
        for merged_range in merged_ranges:
                coord = f'{get_column_letter(merged_range.min_col)}{merged_range.min_row - group["first_row"] + last_header_row}:' \
                        f'{get_column_letter(merged_range.max_col)}{merged_range.max_row - group["first_row"] + last_header_row}'
                new_ws.merge_cells(coord)

        collapse_groups(new_ws, last_header_row)



def main():
    print("Инициализация...")

    # Загрузка конфигурации
    config = configparser.ConfigParser()
    config_path = '.\\config.ini'
    encoding = detect_encoding(config_path)
    config.read(config_path, encoding=encoding)

    src_file = config.get('Settings', 'src_file')
    sheet_name = config.get('Settings', 'sheet')
    tag = config.get('Settings', 'tag')

    if not path.isfile(src_file):
        print(f"Файл {src_file} не найден, проверьте config.ini[src_file] и имя файла")
        input("Нажмите Enter для выхода...")
        return

    print(f"Обработка файла: {src_file}")
    print(f"Исходный лист: {sheet_name}")
    print(f"Тег: {tag}")

    # Загрузка рабочей книги
    workdir = pwd()
    wb = load_workbook(f'{workdir}\\{src_file}', read_only=False)
    ws = wb[sheet_name]
    target_path = f'{workdir}\\groups2sheets-{src_file}'
    # Определение групп
    print("Поиск групп...")
    last_header_row = get_first_group_row(ws)
    groups = get_groups(ws, tag, last_header_row)

    if len(groups) == 0:
        print("Группы не найдены!")
        wb.close()
        input("Нажмите Enter для выхода...")
        return

    print(f"Найдено групп: {len(groups)}")

    # Создание листов с прогресс-баром
    create_group_sheets(wb, groups, ws, last_header_row)
    # Копирование данных с прогресс-баром
    copy_group_data(wb, sheet_name, groups, last_header_row)


    # Сохранение результатов
    print(f"Сохранение результата: {target_path}")
    wb.save(target_path)
    wb.close()

    print("✅ Обработка завершена успешно")
    input("Нажмите Enter для выхода...")


if __name__ == "__main__":
    main()
